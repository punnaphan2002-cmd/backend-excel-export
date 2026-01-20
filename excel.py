import sys
import json
from openpyxl import load_workbook

# argv
# sys.argv[1] = output.xlsx (ไฟล์ที่ copy แล้ว)
# sys.argv[2] = JSON edits

file_path = sys.argv[1]
edits = json.loads(sys.argv[2])

wb = load_workbook(file_path)

for edit in edits:
    sheet_name = edit["sheet"]
    row = edit["row"]        # Excel row (1-based)
    col = edit["col"]        # Excel col (1-based)
    value = edit["value"]

    if sheet_name not in wb.sheetnames:
        continue

    ws = wb[sheet_name]
    ws.cell(row=row, column=col).value = value

wb.save(file_path)
